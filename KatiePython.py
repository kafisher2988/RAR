#Script to modify a spreadsheet
#python katie.py --bookName "" --sheetName ""
#http://openpyxl.readthedocs.io/en/stable/tutorial.html

#Modules
from openpyxl import load_workbook

#Dictionary for making modifications
d = {
	'Escalation_Reason': 'Recommendation',
	'KJR Test 1': 'KJR Test 2',
	'Contact/s cannot be reached' : 'Provide updated business contacts',
	'Extension' : 'Review supplier extension request',
	'Decline: Cost' : 'Escalation Template - does not want to move forward',
	'Document Exemption' : 'Review supplier document exemption'
	}

#open the book
book = load_workbook('Katie.xs\', read_only=False, keep_vba=True)
#book.sheetnames

#open the sheet
sheet = book['RAR_List']
#row
#sheet[1]
#cell
#sheet.cell(1,1).value)

#determine first column (or hardcode?)
#ER_column = [x.value for x in sheet[1]].index('Escalation_Reason') + 1
ER_column = 28

#determine second column (or hardcode?)
#REC_column = [x.value for x in sheet[1]].index('Recommendation') + 1
REC_column = 12

#iterate through and make changes
for rowNum, row in enumerate(sheet):
	ER_value = sheet.cell((rowNum + 1),ER_column).value
	if ER_value in d.keys():
		change = sheet.cell((rowNum + 1),REC_column,d[ER_value])

#Save the book
book.save('Mills.xlsm')
